from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse, HttpResponse
from django.http import QueryDict
from django.contrib import messages
# from .tasks import send_contact_email  # Email sending disabled
from django.conf import settings
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.decorators import login_required
from django.contrib.admin.views.decorators import staff_member_required
from .models import Video, VideoType, Service, Contact, SiteSettings, ServiceIcon
from .forms import ContactForm
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
from django.views.decorators.http import require_http_methods  # ✅ Add this


# from django.core.mail import EmailMultiAlternatives  # Unused after disabling email
# from django.template.loader import render_to_string  # Unused after disabling email
# from django.utils.timezone import now  # Unused here

from django.core.cache import cache
from .models import Video, Service, VideoType, SiteSettings
from .forms import ContactForm

import os, re, uuid
from django.core.paginator import Paginator

def safe_s3_key(folder, filename):
    """Generate a safe, unique S3 key for uploaded files."""
    # Get extension
    _, ext = os.path.splitext(filename)
    ext = ext.lower()

    # Clean base name (remove spaces, special chars)
    base = re.sub(r'[^a-zA-Z0-9_-]', '_', os.path.splitext(filename)[0]).lower()

    # Append unique ID to avoid collisions
    unique_name = f"{base}_{uuid.uuid4().hex}{ext}"

    return f"{folder}/{unique_name}"


def home(request):
    """Optimized Home page view with caching & fewer DB hits"""

    # ✅ Cache site settings (only 1 row, rarely changes)
    site_settings = cache.get("site_settings")
    if site_settings is None:
        site_settings = SiteSettings.objects.first()
        cache.set("site_settings", site_settings, 0)

    # ✅ Cache featured videos (since usually static)
    videos = cache.get("featured_videos")
    if videos is None:
        videos = (
            Video.objects.filter(is_featured=True)
            .select_related("video_type")
            .order_by("order", "-video_created_at")[:6]
        )

        # ✅ Patch attributes so template doesn't break
        for v in videos:
            v.get_video_url = v.get_video_url()
            v.get_thumbnail_url = v.get_thumbnail_url()

        videos = list(videos)
        cache.set("featured_videos", videos, 0)

    # ✅ Cache video types
    video_types = cache.get("video_types")
    if video_types is None:
        video_types = list(VideoType.objects.all())
        cache.set("video_types", video_types, 0)

    # ✅ Services
    services = (
        Service.objects.filter(is_active=True)
        .select_related("icon")
        .order_by("order", "name")
    )

    # Contact form
    form = ContactForm()

    context = {
        "videos": videos,
        "services": services,
        "video_types": video_types,
        "site_settings": site_settings,
        "form": form,
    }
    return render(request, "core/home.html", context)


@csrf_exempt
@require_http_methods(["POST"])
def contact_ajax(request):
    """AJAX endpoint for contact form submission"""
    if request.method != "POST":
        return JsonResponse({'success': False, 'message': 'Invalid request method'}, status=405)

    try:
        # Try to parse JSON data first
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            # If JSON parsing fails, try to get data from POST
            data = request.POST.dict()
        
        print("Received data:", data)  # Debug print
        
        # Prepare a QueryDict to properly handle multi-value fields like services
        qd = QueryDict('', mutable=True)
        for key, value in data.items():
            if key == 'services':
                if isinstance(value, list):
                    qd.setlist('services', [str(v) for v in value])
                elif isinstance(value, str):
                    if ',' in value:
                        qd.setlist('services', [s for s in [x.strip() for x in value.split(',')] if s])
                    elif value:
                        qd.setlist('services', [value])
            elif value is not None:
                qd[key] = str(value)

        form = ContactForm(qd)

        if form.is_valid():
            contact = form.save(commit=False)  # don't save M2M yet
            contact.save()
            form.save_m2m()  # now this works

            # Build email content
            subject = f'New Contact Form Submission - {contact.name}'
            selected_services = ', '.join(contact.services.values_list('name', flat=True)) or 'Not provided'
            message = f"""
New contact form submission received:

Name: {contact.name}
Email: {contact.email}
Phone: {contact.contact}
Business Name: {contact.business_name or 'Not provided'}
Instagram ID: {contact.insta_id or 'Not provided'}
City: {contact.city or 'Not provided'}
Service(s): {selected_services}
Message: {contact.message or 'No message provided'}

Submitted at: {contact.created_at.strftime('%Y-%m-%d %H:%M:%S')}
"""

            # Email sending disabled: details are stored in admin only

            return JsonResponse({
                'success': True,
                'message': 'Thank you for your interest! We will contact you within 24 hours to discuss your premium project.'
            })

        else:
            # Form is not valid
            print("Form errors:", form.errors)  # Debug print
            return JsonResponse({'success': False, 'errors': form.errors})

    except Exception as e:
        print("Error in contact_ajax:", e)
        return JsonResponse({'success': False, 'message': 'An error occurred. Please try again.'})
    
    
def video_detail(request, pk):
    """Video detail view"""
    video = get_object_or_404(Video, pk=pk)
    return JsonResponse({
        'video_name': video.video_name,
        'video_description': video.video_description,
        'video_url': video.get_video_url(),
        'embed_url': video.get_embed_url(),
        'is_youtube': video.is_youtube(),
        'is_vimeo': video.is_vimeo(),
        'is_local_file': video.is_local_file(),
        'thumbnail': video.thumbnail_file.url if video.thumbnail_file else video.thumbnail_url,
    })

def get_videos_by_type(request, type_id):
    videos = Video.objects.filter(video_type_id=type_id)
    video_data = []
    for video in videos:
        video_data.append({
            'id': video.id,
            'video_name': video.video_name,
            'video_description': video.video_description,
            'thumbnail': video.thumbnail_file.url if video.thumbnail_file else video.thumbnail_url,
            'video_url': video.video_file.url if video.video_file else video.video_url,
        })
    return JsonResponse({'videos': video_data})

# Custom Admin Panel Views
from django.db.models import Count, Q

@staff_member_required
def custom_admin_dashboard(request):
    total_videos = Video.objects.count()
    total_contacts = Contact.objects.count()
    uncontacted_leads = Contact.objects.filter(is_contacted=False).count()
    total_services = Service.objects.count()

    recent_contacts = Contact.objects.order_by('-created_at')[:5]
    recent_videos = Video.objects.select_related('video_type').order_by('-video_created_at')[:5]

    context = {
        'total_videos': total_videos,
        'total_contacts': total_contacts,
        'uncontacted_leads': uncontacted_leads,
        'total_services': total_services,
        'recent_contacts': recent_contacts,
        'recent_videos': recent_videos,
    }
    return render(request, 'admin/custom_dashboard.html', context)


@staff_member_required
def custom_admin_videos(request):
    videos = Video.objects.filter(is_featured=True).select_related('video_type').order_by('order', '-video_created_at')
    video_types = VideoType.objects.all()
    context = {
        'videos': videos,
        'video_types': video_types,
    }
    return render(request, 'admin/videos.html', context)

@staff_member_required
def custom_admin_contacts(request):
    contacts_qs = Contact.objects.prefetch_related('services').all().order_by('-created_at')

    q = request.GET.get('q', '').strip()
    contacted = request.GET.get('contacted', '').strip()
    service_id = request.GET.get('service', '').strip()

    if q:
        from django.db.models import Q
        contacts_qs = contacts_qs.filter(
            Q(name__icontains=q) |
            Q(email__icontains=q) |
            Q(contact__icontains=q) |
            Q(business_name__icontains=q) |
            Q(insta_id__icontains=q) |
            Q(city__icontains=q) |
            Q(message__icontains=q)
        )

    if contacted == 'yes':
        contacts_qs = contacts_qs.filter(is_contacted=True)
    elif contacted == 'no':
        contacts_qs = contacts_qs.filter(is_contacted=False)

    if service_id.isdigit():
        contacts_qs = contacts_qs.filter(services__id=int(service_id))

    # Pagination
    paginator = Paginator(contacts_qs, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    services = Service.objects.all()  # no extra ORDER BY

    # Only one count now
    filtered_count = paginator.count

    context = {
        'contacts': page_obj.object_list,
        'page_obj': page_obj,
        'paginator': paginator,
        'services': services,
        'q': q,
        'selected_service': service_id,
        'selected_contacted': contacted,
        'filtered_count': filtered_count,
    }
    return render(request, 'admin/contacts.html', context)


@staff_member_required
def custom_admin_services(request):
    services = Service.objects.all().order_by('order')
    icons = ServiceIcon.objects.all()
    context = {
        'services': services,
        'icons': icons,
    }
    return render(request, 'admin/services.html', context)

@staff_member_required
def export_contacts_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Contact Submissions"

    headers = [
        'Name', 'Email', 'Contact', 'Business Name', 'Instagram ID',
        'Services', 'City', 'Message', 'Submitted Date', 'Contacted', 'Notes'
    ]

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    contacts = Contact.objects.all().order_by('-created_at')
    for row, contact in enumerate(contacts, 2):
        ws.cell(row=row, column=1,  value=contact.name)
        ws.cell(row=row, column=2,  value=contact.email)
        ws.cell(row=row, column=3,  value=contact.contact)
        ws.cell(row=row, column=4,  value=contact.business_name)
        ws.cell(row=row, column=5,  value=contact.insta_id)
        ws.cell(row=row, column=6,  value=", ".join(contact.services.values_list('name', flat=True)))
        ws.cell(row=row, column=7,  value=contact.city)
        ws.cell(row=row, column=8,  value=contact.message)
        ws.cell(row=row, column=9,  value=contact.created_at.strftime('%Y-%m-%d %H:%M'))
        ws.cell(row=row, column=10, value="Yes" if contact.is_contacted else "No")
        ws.cell(row=row, column=11, value=contact.notes)

    # autosize columns (unchanged) ...
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=contacts_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    wb.save(response)
    return response


import boto3
from django.conf import settings
from django.views.decorators.http import require_GET

@require_GET
@login_required
def generate_presigned_url(request):
    """Generate presigned URL for direct S3 upload (video or thumbnail)"""
    file_name = request.GET.get("file_name")
    file_type = request.GET.get("file_type", "video/mp4")
    folder = request.GET.get("folder", "videos")   # default to videos, but can be 'thumbnails'

    if not file_name:
        return JsonResponse({"error": "Missing file_name"}, status=400)

    s3 = boto3.client(
        "s3",
        aws_access_key_id=settings.AWS_ACCESS_KEY_ID,
        aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY,
        region_name=settings.AWS_S3_REGION_NAME
    )

    # ✅ Sanitize & normalize filename before upload
    key = safe_s3_key(folder, file_name)

    # Presigned PUT URL for upload
    presigned_url = s3.generate_presigned_url(
        "put_object",
        Params={
            "Bucket": settings.AWS_STORAGE_BUCKET_NAME,
            "Key": key,
            "ContentType": file_type,
        },
        ExpiresIn=3600
    )

    # ✅ Region-specific permanent file URL
    file_url = (
        f"https://{settings.AWS_STORAGE_BUCKET_NAME}.s3."
        f"{settings.AWS_S3_REGION_NAME}.amazonaws.com/{key}"
    )

    return JsonResponse({
        "upload_url": presigned_url,  # frontend uses this to upload
        "file_url": file_url          # save this in DB
    })




from django.shortcuts import render

def error_page(request, code, title, message, status):
    return render(request, "admin/error.html", {
        "code": code,
        "title": title,
        "message": message
    }, status=status)

def custom_404(request, exception):
    return error_page(request, 404, "Page Not Found", "Sorry, the page you’re looking for doesn’t exist.", 404)

def custom_500(request):
    return error_page(request, 500, "Server Error", "Something went wrong on our end. Please try again later.", 500)

def custom_403(request, exception):
    return error_page(request, 403, "Forbidden", "You don’t have permission to access this page.", 403)
