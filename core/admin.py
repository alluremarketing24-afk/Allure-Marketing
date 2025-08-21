from django.contrib import admin
from django.http import HttpResponse
from django.urls import path
from django.contrib.admin import AdminSite
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from .models import Service, Video, VideoType, Contact, SiteSettings, ServiceIcon
import datetime
from adminsortable2.admin import SortableAdminMixin


class AllureMarketingAdminSite(AdminSite):
    site_header = "Allure Marketing Admin"
    site_title = "Allure Marketing"
    index_title = "Welcome to Allure Marketing Administration"

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('export-contacts/', self.admin_view(self.export_contacts), name='export_contacts'),
        ]
        return custom_urls + urls

    def export_contacts(self, request):
        wb = Workbook()
        ws = wb.active
        ws.title = "Contacts"

        headers = [
            'Name', 'Contact', 'Business Name', 'Instagram ID', 'City',
            'Decision Maker', 'Email', 'Message', 'Created At', 'Contacted', 'Notes'
        ]

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        contacts = Contact.objects.all().order_by('-created_at')
        for row, contact in enumerate(contacts, 2):
            ws.cell(row=row, column=1, value=contact.name)
            ws.cell(row=row, column=2, value=contact.contact)
            ws.cell(row=row, column=3, value=contact.business_name)
            ws.cell(row=row, column=4, value=contact.insta_id)
            ws.cell(row=row, column=5, value=contact.city)
            ws.cell(row=row, column=6, value=contact.get_is_decision_maker_display())
            ws.cell(row=row, column=7, value=contact.email)
            ws.cell(row=row, column=8, value=contact.message)
            ws.cell(row=row, column=9, value=contact.created_at.strftime('%Y-%m-%d %H:%M:%S'))
            ws.cell(row=row, column=10, value="Yes" if contact.is_contacted else "No")
            ws.cell(row=row, column=11, value=contact.notes)

        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=contacts_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

        wb.save(response)
        return response


# Create custom admin site instance
admin_site = AllureMarketingAdminSite(name='allure_admin')


@admin.register(Service, site=admin_site)
class ServiceAdmin(admin.ModelAdmin):
    list_display = ['name', 'get_icon_name', 'is_active', 'order', 'created_at']
    list_filter = ['is_active', 'created_at']
    search_fields = ['name', 'description']
    list_editable = ['is_active', 'order']
    ordering = ['order', 'name']

    fieldsets = (
        ('Basic Information', {
            'fields': ('name', 'description', 'icon')
        }),
        ('Key Points', {
            'fields': (
                'key_point_1', 'key_point_2', 'key_point_3',
                'key_point_4', 'key_point_5', 'key_point_6'
            )
        }),
        ('Display Settings', {
            'fields': ('is_active', 'order')
        })
    )

    def get_icon_name(self, obj):
        return obj.icon.name if obj.icon else '-'
    get_icon_name.short_description = 'Icon'


@admin.register(VideoType, site=admin_site)
class VideoTypeAdmin(admin.ModelAdmin):
    list_display = ['name', 'description']
    search_fields = ['name']


@admin.register(Video, site=admin_site)
class VideoAdmin(admin.ModelAdmin):
    list_display = ['video_name', 'video_type', 'is_featured', 'order', 'video_created_at']
    list_filter = ['video_type', 'is_featured', 'video_created_at']
    search_fields = ['video_name', 'video_description']
    list_editable = ['is_featured', 'order']
    ordering = ['-video_created_at']
    readonly_fields = ['video_url', 'thumbnail_url']

    fields = [
        'video_name', 'video_type', 'video_description',
        'video_file', 'thumbnail_file',   # ✅ S3 upload fields
        'video_url', 'thumbnail_url',
        'is_featured', 'order'
    ]

    def save_model(self, request, obj, form, change):
        """
        Ensure S3 URLs are saved after upload.
        """
        if obj.video_file:
            obj.video_url = obj.video_file.url
        if obj.thumbnail_file:
            obj.thumbnail_url = obj.thumbnail_file.url
        super().save_model(request, obj, form, change)


@admin.register(Contact, site=admin_site)
class ContactAdmin(admin.ModelAdmin):
    list_display = ['name', 'email', 'contact', 'business_name', 'display_services', 'is_contacted', 'created_at']
    list_filter = ['is_contacted', 'created_at', 'city', 'services']
    search_fields = ['name', 'email', 'business_name', 'contact']
    list_editable = ['is_contacted']
    ordering = ['-created_at']
    readonly_fields = ['created_at']

    actions = ['mark_as_contacted', 'mark_as_not_contacted', 'export_selected_contacts']

    def mark_as_contacted(self, request, queryset):
        updated = queryset.update(is_contacted=True)
        self.message_user(request, f'{updated} contacts marked as contacted.')

    def mark_as_not_contacted(self, request, queryset):
        updated = queryset.update(is_contacted=False)
        self.message_user(request, f'{updated} contacts marked as not contacted.')

    def export_selected_contacts(self, request, queryset):
        wb = Workbook()
        ws = wb.active
        ws.title = "Selected Contacts"

        headers = [
            'Name', 'Contact', 'Business Name', 'Instagram ID', 'City',
            'Services', 'Email', 'Message', 'Created At', 'Contacted', 'Notes'
        ]

        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        for row, contact in enumerate(queryset.order_by('-created_at'), 2):
            ws.cell(row=row, column=1, value=contact.name)
            ws.cell(row=row, column=2, value=contact.contact)
            ws.cell(row=row, column=3, value=contact.business_name)
            ws.cell(row=row, column=4, value=contact.insta_id)
            ws.cell(row=row, column=5, value=contact.city)
            ws.cell(row=row, column=6, value=", ".join(contact.services.values_list('name', flat=True)))
            ws.cell(row=row, column=7, value=contact.email)
            ws.cell(row=row, column=8, value=contact.message)
            ws.cell(row=row, column=9, value=contact.created_at.strftime('%Y-%m-%d %H:%M:%S'))
            ws.cell(row=row, column=10, value="Yes" if contact.is_contacted else "No")
            ws.cell(row=row, column=11, value=contact.notes)

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=selected_contacts_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

        wb.save(response)
        return response

    def display_services(self, obj):
        return ", ".join(obj.services.values_list('name', flat=True))
    display_services.short_description = 'Services'
    export_selected_contacts.short_description = "Export selected contacts to Excel"


@admin.register(SiteSettings, site=admin_site)
class SiteSettingsAdmin(admin.ModelAdmin):
    fieldsets = (
        ('Basic Information', {
            'fields': ('site_name', 'tagline', 'meta_description', 'meta_keywords')
        }),
        ('Contact Information', {
            'fields': ('contact_email', 'contact_phone', 'whatsapp_number', 'address')
        }),
        ('Homepage Content', {
            'fields': ('hero_title', 'hero_subtitle', 'logo')
        }),
        ('About Section', {
            'fields': ('about_text',)
        }),
    )

    def has_add_permission(self, request):
        return not SiteSettings.objects.exists()

    def has_delete_permission(self, request, obj=None):
        return False


# Register custom admin site
admin.site = admin_site
